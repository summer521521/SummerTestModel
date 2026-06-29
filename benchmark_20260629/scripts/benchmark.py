from __future__ import annotations

import argparse
import ast
import csv
import datetime as dt
import json
import math
import re
import subprocess
import sys
import tempfile
import textwrap
import time
import urllib.error
import urllib.request
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Iterable


OLLAMA_URL = "http://127.0.0.1:11434/api/generate"
TAGS_URL = "http://127.0.0.1:11434/api/tags"


@dataclass(frozen=True)
class TestCase:
    id: str
    name: str
    category: str
    max_score: int
    prompt: str
    grader: Callable[[str], tuple[int, str]]


def first_json_object(text: str) -> dict | None:
    text = text.strip()
    candidates = []
    if text.startswith("{"):
        candidates.append(text)
    candidates.extend(re.findall(r"\{.*?\}", text, flags=re.S))
    for candidate in candidates:
        try:
            value = json.loads(candidate)
            if isinstance(value, dict):
                return value
        except json.JSONDecodeError:
            continue
    return None


def normalize(text: str) -> str:
    return re.sub(r"\s+", "", text.lower())


def grade_format_json(text: str) -> tuple[int, str]:
    obj = first_json_object(text)
    if obj is None:
        return 0, "没有输出可解析 JSON"
    score = 0
    notes = []
    if set(["total", "valid", "code", "summary"]).issubset(obj.keys()):
        score += 2
    else:
        notes.append("字段不全")
    try:
        if abs(float(obj.get("total")) - 164.7) < 0.011:
            score += 3
        else:
            notes.append("total 错")
    except Exception:
        notes.append("total 不是数字")
    if obj.get("valid") is True:
        score += 2
    else:
        notes.append("valid 错")
    if obj.get("code") == "OK-16470":
        score += 2
    else:
        notes.append("code 错")
    extra = text.strip()
    if extra.startswith("{") and extra.endswith("}"):
        score += 1
    else:
        notes.append("JSON 外有额外文本")
    return score, "；".join(notes) if notes else "完全正确"


def grade_math_reasoning(text: str) -> tuple[int, str]:
    obj = first_json_object(text)
    target_time = "10:00"
    target_count = 43
    score = 0
    notes = []
    if obj:
        time_value = str(obj.get("time", ""))
        count_value = obj.get("count")
        if target_time in time_value:
            score += 5
        else:
            notes.append("时间错")
        try:
            if int(count_value) == target_count:
                score += 5
            else:
                notes.append("次数错")
        except Exception:
            notes.append("次数不可解析")
        return score, "；".join(notes) if notes else "完全正确"
    compact = normalize(text)
    if "10:00" in compact:
        score += 5
    else:
        notes.append("时间错")
    if re.search(r"(?<!\d)43(?!\d)", text):
        score += 5
    else:
        notes.append("次数错")
    return score, "；".join(notes) if notes else "完全正确"


def grade_long_context(text: str) -> tuple[int, str]:
    compact = normalize(text)
    score = 0
    notes = []
    if "helios" in compact:
        score += 6
    else:
        notes.append("目标项目错")
    if "lin" in compact:
        score += 1
    if "paused" in compact or "暂停" in compact:
        score += 1
    if "apac" not in compact or "非apac" in compact or "不是apac" in compact:
        score += 1
    if "orion" not in compact and "atlas" not in compact and "nova" not in compact:
        score += 1
    else:
        notes.append("混入干扰项")
    return min(score, 10), "；".join(notes) if notes else "正确定位"


def grade_translation(text: str) -> tuple[int, str]:
    score = 0
    notes = []
    required_terms = ["RAG", "KV cache", "LoRA", "p95 latency", "tenant_id"]
    for term in required_terms:
        if term in text:
            score += 1
        else:
            notes.append(f"丢失 {term}")
    if not re.search(r"[\u4e00-\u9fff]", text):
        score += 2
    else:
        notes.append("仍含中文")
    lowered = text.lower()
    if "do not" in lowered or "must not" in lowered or "cannot" in lowered:
        score += 1
    else:
        notes.append("约束语气不足")
    if "tenant" in lowered and ("isolation" in lowered or "isolated" in lowered):
        score += 1
    else:
        notes.append("租户隔离表达弱")
    if 35 <= len(text.split()) <= 120:
        score += 1
    else:
        notes.append("长度异常")
    return min(score, 10), "；".join(notes) if notes else "术语和约束保留良好"


def grade_hallucination(text: str) -> tuple[int, str]:
    compact = normalize(text)
    score = 0
    notes = []
    uncertainty_hits = ["无法确定", "不能确定", "缺少", "不足", "unknown", "cannot determine", "not enough"]
    if any(hit in compact for hit in uncertainty_hits):
        score += 5
    else:
        notes.append("没有承认信息不足")
    if "2024" in compact and "黑曜计划" in text:
        score += 2
    else:
        notes.append("没有保留已知事实")
    fabricated = ["张", "王", "李", "ceo", "创始人", "北京", "上海", "美国", "融资"]
    if not any(x in compact for x in fabricated):
        score += 3
    else:
        notes.append("疑似编造细节")
    return min(score, 10), "；".join(notes) if notes else "拒绝编造且保留已知信息"


def extract_code(text: str) -> str:
    match = re.search(r"```(?:python)?\s*(.*?)```", text, flags=re.S | re.I)
    if match:
        return match.group(1).strip()
    return text.strip()


def grade_code_bugfix(text: str) -> tuple[int, str]:
    code = extract_code(text)
    if "def top_k_words" not in code:
        return 0, "没有给出目标函数"
    try:
        ast.parse(code)
    except SyntaxError as exc:
        return 1, f"语法错误：{exc.msg}"

    harness = textwrap.dedent(
        """
        import json
        ns = {}
        exec(CODE, ns)
        fn = ns.get("top_k_words")
        assert fn is not None
        cases = [
            ("Apple banana apple, BANANA! pear.", 2, [("apple", 2), ("banana", 2)]),
            ("b a c b c c", 3, [("c", 3), ("b", 2), ("a", 1)]),
            ("Hi... hi? AI ai; ai", 2, [("ai", 3), ("hi", 2)]),
        ]
        for text, k, expected in cases:
            got = fn(text, k)
            assert got == expected, (text, got, expected)
        print("PASS")
        """
    )
    with tempfile.TemporaryDirectory() as tmp:
        runner = Path(tmp) / "run_code_test.py"
        runner.write_text("CODE = " + repr(code) + "\n" + harness, encoding="utf-8")
        try:
            result = subprocess.run(
                [sys.executable, str(runner)],
                capture_output=True,
                text=True,
                timeout=5,
            )
        except subprocess.TimeoutExpired:
            return 2, "代码执行超时"
    if result.returncode == 0 and "PASS" in result.stdout:
        return 10, "单元测试全通过"
    return 4, (result.stderr or result.stdout).strip()[:240]


def grade_planning(text: str) -> tuple[int, str]:
    compact = normalize(text)
    score = 0
    notes = []
    if "有可行解" in text or "是否有可行解：是" in text or "可行解：是" in text:
        score += 2
    else:
        notes.append("没有明确有解")
    if "07:00" in compact and "08:00" in compact and ("08:30" in compact or "恢复" in text):
        score += 2
    else:
        notes.append("A/恢复安排不清")
    if "10:45" in compact and "12:30" in compact:
        score += 2
    else:
        notes.append("C 在 D 前不清")
    if "11:45" in compact and "12:00" in compact and ("等待" in text or "不在路上" in text or "餐馆" in text):
        score += 2
    else:
        notes.append("12:00-13:00 禁行处理不清")
    if "15:55" in compact or "16:45" in compact or "17:00" in compact:
        score += 1
    else:
        notes.append("F 完成时间不清")
    if "B" in text and "F" in text and ("A" in text or "跑步" in text):
        score += 1
    else:
        notes.append("关键任务缺失")
    return min(score, 10), "；".join(notes) if notes else "关键约束基本覆盖"


TESTS = [
    TestCase(
        id="format_json",
        name="JSON 格式与算术",
        category="格式遵循",
        max_score=10,
        grader=grade_format_json,
        prompt=(
            "你是严谨的数据处理器。只输出一个 JSON 对象，不要 Markdown，不要解释。\n"
            "订单金额：A=128.50，B=73.25，退款=-18.75。先求小计，再打 9 折。\n"
            "输出字段：total（最终金额，保留 2 位小数的数字）、valid（布尔值）、code、summary。\n"
            "code 必须是 OK- 加上最终金额乘以 100 后的整数。"
        ),
    ),
    TestCase(
        id="math_reasoning",
        name="多步数学推理",
        category="推理",
        max_score=10,
        grader=grade_math_reasoning,
        prompt=(
            "三台设备从 08:00 同时启动。A 每 6 分钟响一次，B 每 8 分钟响一次，C 每 15 分钟响一次。\n"
            "问题：08:00 之后第一次三台再次同时响铃是什么时间？从 08:00 之后到这次同时响铃为止，"
            "如果每台设备每响一次都单独计数，总响铃次数是多少？\n"
            "只输出 JSON：{\"time\":\"HH:MM\",\"count\":数字}"
        ),
    ),
    TestCase(
        id="long_context",
        name="长上下文检索",
        category="长上下文",
        max_score=10,
        grader=grade_long_context,
        prompt=(
            "下面是项目台账。请找出唯一满足条件的项目：status 是 PAUSED，owner 是 Lin，region 不是 APAC。"
            "只输出项目名和一句理由。\n"
            "1. Atlas | owner=Chen | region=APAC | status=ACTIVE\n"
            "2. Boreal | owner=Lin | region=APAC | status=PAUSED\n"
            "3. Cinder | owner=Mei | region=EU | status=PAUSED\n"
            "4. Delta | owner=Lin | region=NA | status=ACTIVE\n"
            "5. Ember | owner=Rao | region=EU | status=ACTIVE\n"
            "6. Fjord | owner=Lin | region=APAC | status=ACTIVE\n"
            "7. Helios | owner=Lin | region=EU | status=PAUSED\n"
            "8. Ion | owner=Qiu | region=NA | status=PAUSED\n"
            "9. Juno | owner=Lin | region=LATAM | status=DONE\n"
            "10. Kestrel | owner=An | region=EU | status=PAUSED\n"
            "11. Luna | owner=Lin | region=APAC | status=PAUSED\n"
            "12. Meridian | owner=Sun | region=NA | status=ACTIVE\n"
            "13. Nova | owner=Lin | region=EU | status=ACTIVE\n"
            "14. Orion | owner=Lin | region=NA | status=PAUSED_BUT_ARCHIVED\n"
            "15. Pulse | owner=Zed | region=EU | status=PAUSED\n"
            "16. Quartz | owner=Lin | region=APAC | status=DONE"
        ),
    ),
    TestCase(
        id="translation_terms",
        name="技术翻译与术语保留",
        category="翻译",
        max_score=10,
        grader=grade_translation,
        prompt=(
            "把下面中文翻译成自然、准确的英文。必须原样保留术语 RAG、KV cache、LoRA、p95 latency、tenant_id，"
            "不要添加原文没有的信息。\n"
            "原文：为了降低 p95 latency，我们把 RAG 检索结果缓存 10 分钟，但不能跨 tenant_id 复用。"
            "微调只允许使用 LoRA，线上推理必须监控 KV cache 命中率。"
        ),
    ),
    TestCase(
        id="anti_hallucination",
        name="抗幻觉与信息边界",
        category="可靠性",
        max_score=10,
        grader=grade_hallucination,
        prompt=(
            "已知事实只有一句：黑曜计划在 2024 年启动。"
            "请回答：黑曜计划的负责人是谁、预算是多少、总部在哪？"
            "要求：不能编造；如果信息不足，请明确指出哪些无法确定。"
        ),
    ),
    TestCase(
        id="code_bugfix",
        name="代码修复",
        category="代码",
        max_score=10,
        grader=grade_code_bugfix,
        prompt=(
            "修复下面 Python 函数。要求：统计单词频次，忽略大小写，去掉英文标点；返回前 k 个"
            "(word, count)，按 count 降序、count 相同按 word 升序。只输出完整 Python 代码。\n\n"
            "def top_k_words(text, k):\n"
            "    counts = {}\n"
            "    for word in text.split(' '):\n"
            "        counts[word] = counts.get(word, 0) + 1\n"
            "    return sorted(counts.items())[:k]\n"
        ),
    ),
    TestCase(
        id="planning_schedule",
        name="多约束日程规划",
        category="规划",
        max_score=10,
        grader=grade_planning,
        prompt=(
            "你是排班助手。为同一个人安排周六一天，必须满足："
            "A 跑步 60 分钟，必须 07:00-10:00 之间开始，结束后连续 30 分钟洗澡恢复，恢复时不能做别的事；"
            "B 买菜 40 分钟，必须在超市 09:30-18:00 内完整完成；"
            "C 写作业 120 分钟，必须拆成两个连续 60 分钟块，中间不能插入别的任务，只能在家；"
            "D 和朋友吃饭固定 12:30 开始，90 分钟，餐馆；"
            "E 图书馆借书 30 分钟，必须 14:00-17:00 内开始；"
            "F 整理房间 45 分钟，只能在家，20:00 前完成。"
            "通勤：家-操场15，家-超市20，家-餐馆25，家-图书馆30，超市-餐馆15，超市-图书馆20，餐馆-图书馆10，操场-超市25，操场-餐馆35，操场-图书馆40。"
            "从家出发，最后回家。12:00-13:00 不能在路上。C 必须在 D 前完成，B 必须在 F 前完成，A 和 F 不能相邻。"
            "按三部分输出：结论、日程表、说明。"
        ),
    ),
]


def safe_name(model: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", model.replace(":", "__").replace("/", "_"))


def get_models() -> list[dict]:
    with urllib.request.urlopen(TAGS_URL, timeout=10) as response:
        data = json.loads(response.read().decode("utf-8"))
    return data.get("models", [])


def call_model(model: str, prompt: str, timeout: int) -> tuple[str, float, str | None]:
    payload = {
        "model": model,
        "prompt": prompt,
        "stream": False,
        "options": {
            "temperature": 0,
            "num_predict": 900,
        },
    }
    req = urllib.request.Request(
        OLLAMA_URL,
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json"},
    )
    started = time.perf_counter()
    try:
        with urllib.request.urlopen(req, timeout=timeout) as response:
            data = json.loads(response.read().decode("utf-8", errors="replace"))
        elapsed = time.perf_counter() - started
        return data.get("response", ""), elapsed, None
    except (urllib.error.URLError, TimeoutError, json.JSONDecodeError) as exc:
        elapsed = time.perf_counter() - started
        return "", elapsed, f"{type(exc).__name__}: {exc}"


def write_suite_readme(run_dir: Path, tests: Iterable[TestCase], models: list[str]) -> None:
    lines = [
        "# Benchmark 20260629",
        "",
        "## 测试内容",
        "",
        "本批次用于评测 Ollama 本地与云端模型在多用途场景下的基础能力：格式遵循、数学推理、长上下文检索、技术翻译、抗幻觉、代码修复和多约束规划。",
        "",
        "## 模型范围",
        "",
    ]
    lines.extend(f"- `{model}`" for model in models)
    lines.extend(["", "## 测试题", ""])
    for test in tests:
        lines.extend([f"### {test.id}：{test.name}", "", f"- 类别：{test.category}", f"- 满分：{test.max_score}", "", "```text", test.prompt, "```", ""])
    (run_dir / "test_suite.md").write_text("\n".join(lines), encoding="utf-8")


def write_csv(path: Path, rows: list[dict]) -> None:
    fieldnames = ["model", "test_id", "category", "score", "max_score", "elapsed_sec", "error", "note"]
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in fieldnames})


def write_xlsx(path: Path, rows: list[dict], model_totals: list[dict]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except Exception:
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Scores"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    headers = ["model", "total", "max_total", "percent", "rank", "failed_tests", "avg_elapsed_sec"]
    for c, value in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=value)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    for r, row in enumerate(model_totals, 2):
        for c, key in enumerate(headers, 1):
            cell = ws.cell(row=r, column=c, value=row.get(key))
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if key != "model" else "left")
    for idx, width in enumerate([34, 10, 10, 10, 8, 12, 16], 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    detail = wb.create_sheet("Details")
    detail_headers = ["model", "test_id", "category", "score", "max_score", "elapsed_sec", "error", "note"]
    for c, value in enumerate(detail_headers, 1):
        cell = detail.cell(row=1, column=c, value=value)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    for r, row in enumerate(rows, 2):
        for c, key in enumerate(detail_headers, 1):
            cell = detail.cell(row=r, column=c, value=row.get(key))
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for idx, width in enumerate([34, 22, 12, 10, 10, 14, 26, 54], 1):
        detail.column_dimensions[get_column_letter(idx)].width = width
    wb.save(path)


def make_report(run_dir: Path, rows: list[dict], tests: list[TestCase], models_meta: dict[str, dict]) -> list[dict]:
    by_model: dict[str, list[dict]] = {}
    for row in rows:
        by_model.setdefault(row["model"], []).append(row)
    max_total = sum(test.max_score for test in tests)
    totals = []
    for model, model_rows in by_model.items():
        total = sum(int(row["score"]) for row in model_rows)
        failed = sum(1 for row in model_rows if row.get("error"))
        elapsed_values = [float(row["elapsed_sec"]) for row in model_rows if row.get("elapsed_sec") is not None]
        totals.append(
            {
                "model": model,
                "total": total,
                "max_total": max_total,
                "percent": round(total / max_total * 100, 1) if max_total else 0,
                "failed_tests": failed,
                "avg_elapsed_sec": round(sum(elapsed_values) / len(elapsed_values), 2) if elapsed_values else 0,
            }
        )
    totals.sort(key=lambda item: (-item["total"], item["failed_tests"], item["avg_elapsed_sec"], item["model"]))
    for idx, item in enumerate(totals, 1):
        item["rank"] = idx

    category_names = sorted({test.category for test in tests})
    lines = [
        "# Benchmark 20260629 结果报告",
        "",
        "## 测试内容",
        "",
        "本批次覆盖格式遵循、数学推理、长上下文检索、技术翻译、抗幻觉、代码修复和多约束规划。每题 10 分，总分 70 分。",
        "",
        "## 过程",
        "",
        "- 使用 Ollama 本地 API `/api/generate` 调用模型。",
        "- 温度设为 0，限制输出长度为 900 token，尽量减少随机性。",
        "- 每个模型每道题保存原始回答到 `raw/`，再用脚本自动判分。",
        "- `qwen3-embedding` 属于嵌入模型，默认不参与生成式问答评测。",
        "",
        "## 总排名",
        "",
        "| 排名 | 模型 | 总分 | 百分比 | 失败题数 | 平均耗时(s) |",
        "| ---: | --- | ---: | ---: | ---: | ---: |",
    ]
    for item in totals:
        lines.append(
            f"| {item['rank']} | `{item['model']}` | {item['total']}/{item['max_total']} | {item['percent']}% | {item['failed_tests']} | {item['avg_elapsed_sec']} |"
        )
    lines.extend(["", "## 分项结果", ""])
    for category in category_names:
        lines.extend([f"### {category}", "", "| 模型 | 测试 | 得分 | 说明 |", "| --- | --- | ---: | --- |"])
        category_rows = [row for row in rows if row["category"] == category]
        category_rows.sort(key=lambda row: (row["test_id"], -int(row["score"]), row["model"]))
        for row in category_rows:
            note = row.get("note") or row.get("error") or ""
            lines.append(f"| `{row['model']}` | {row['test_id']} | {row['score']}/{row['max_score']} | {note.replace('|', '/')} |")
        lines.append("")

    lines.extend(["## 模型概览", "", "| 模型 | 类型 | 大小 | 备注 |", "| --- | --- | ---: | --- |"])
    for item in totals:
        model = item["model"]
        meta = models_meta.get(model, {})
        size = meta.get("size")
        size_text = "-" if not size else f"{round(size / (1024**3), 2)} GB"
        kind = "云端" if model.endswith(":cloud") or "-cloud" in model else "本地"
        note = "嵌入模型已排除" if "embedding" in model.lower() else ""
        lines.append(f"| `{model}` | {kind} | {size_text} | {note} |")

    (run_dir / "report.md").write_text("\n".join(lines), encoding="utf-8")
    return totals


def main() -> int:
    parser = argparse.ArgumentParser(description="Run Ollama model benchmark.")
    parser.add_argument("--models", nargs="*", help="Specific model names. Defaults to all non-embedding models.")
    parser.add_argument("--tests", nargs="*", help="Specific test ids. Defaults to all tests.")
    parser.add_argument("--timeout", type=int, default=180, help="Per model/test timeout in seconds.")
    parser.add_argument("--output", default=None, help="Run output directory.")
    parser.add_argument("--skip-cloud", action="store_true", help="Skip models whose name looks cloud-hosted.")
    parser.add_argument("--limit", type=int, default=None, help="Limit number of models for smoke tests.")
    args = parser.parse_args()

    available = get_models()
    meta = {model["name"]: model for model in available}
    if args.models:
        models = args.models
    else:
        models = [model["name"] for model in available if "embedding" not in model["name"].lower()]
    if args.skip_cloud:
        models = [model for model in models if not (model.endswith(":cloud") or "-cloud" in model)]
    if args.limit is not None:
        models = models[: args.limit]

    selected_tests = TESTS
    if args.tests:
        wanted = set(args.tests)
        selected_tests = [test for test in TESTS if test.id in wanted]
    if not selected_tests:
        raise SystemExit("No tests selected.")

    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    run_dir = Path(args.output or f"benchmark_20260629/runs/{stamp}")
    raw_dir = run_dir / "raw"
    raw_dir.mkdir(parents=True, exist_ok=True)
    write_suite_readme(run_dir, selected_tests, models)

    rows: list[dict] = []
    total_jobs = len(models) * len(selected_tests)
    job_idx = 0
    for model in models:
        model_dir = raw_dir / safe_name(model)
        model_dir.mkdir(parents=True, exist_ok=True)
        for test in selected_tests:
            job_idx += 1
            print(f"[{job_idx}/{total_jobs}] {model} :: {test.id}", flush=True)
            response, elapsed, error = call_model(model, test.prompt, args.timeout)
            (model_dir / f"{test.id}.txt").write_text(response, encoding="utf-8")
            if error:
                score, note = 0, error
            else:
                score, note = test.grader(response)
            row = {
                "model": model,
                "test_id": test.id,
                "category": test.category,
                "score": score,
                "max_score": test.max_score,
                "elapsed_sec": round(elapsed, 2),
                "error": error or "",
                "note": note,
            }
            rows.append(row)
            write_csv(run_dir / "scores.csv", rows)
            totals = make_report(run_dir, rows, selected_tests, meta)
            write_xlsx(run_dir / "scores.xlsx", rows, totals)

    totals = make_report(run_dir, rows, selected_tests, meta)
    write_csv(run_dir / "scores.csv", rows)
    write_xlsx(run_dir / "scores.xlsx", rows, totals)
    print(f"Report: {run_dir / 'report.md'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
